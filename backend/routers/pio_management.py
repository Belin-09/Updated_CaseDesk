import os
import re
import tempfile
import openpyxl
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from sqlalchemy.orm import Session
from database import get_db
from models import Case, ConfirmedPIO, AuditLog
from auth import get_current_user
from extractor.detect import extract_text, detect_file_type

router = APIRouter(prefix="/pio", tags=["PIO Intelligence"])

def extract_numbers_from_text(text: str) -> list[str]:
    # Remove spaces and dashes to easily match continuous digits
    clean_text = text.replace(" ", "").replace("-", "")
    # Look for numbers between 10 and 15 digits, optionally starting with +
    pattern = re.compile(r'\+?\d{10,15}')
    matches = pattern.findall(clean_text)
    # Strip leading '+' from all matches so we store pure digits
    return list(set([m.lstrip('+') for m in matches]))

@router.post("/upload-master-list")
async def upload_master_list(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in {".pdf", ".docx", ".xlsx"}:
        raise HTTPException(status_code=400, detail="Only .pdf, .docx, and .xlsx files are supported")
        
    temp_dir = tempfile.gettempdir()
    temp_path = os.path.join(temp_dir, file.filename)
    
    try:
        content = await file.read()
        with open(temp_path, "wb") as f:
            f.write(content)
            
        all_text = ""
        if ext == ".xlsx":
            wb = openpyxl.load_workbook(temp_path, data_only=True)
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                for row in ws.iter_rows(values_only=True):
                    row_text = " ".join([str(cell) for cell in row if cell is not None])
                    all_text += row_text + " "
        else:
            file_type = detect_file_type(temp_path)
            all_text, _ = extract_text(temp_path, file_type)
            
        numbers = extract_numbers_from_text(all_text)
        
        # Insert new numbers into DB
        added_count = 0
        for num in numbers:
            exists = db.query(ConfirmedPIO).filter(ConfirmedPIO.phone_number == num).first()
            if not exists:
                db.add(ConfirmedPIO(phone_number=num))
                added_count += 1
                
        db.commit()
        
        total_count = db.query(ConfirmedPIO).count()
        
        log = AuditLog(
            username=current_user.username,
            action="UPLOAD_MASTER_PIO_LIST",
            details=f"Uploaded {file.filename}, extracted {added_count} new numbers. Total confirmed numbers: {total_count}"
        )
        db.add(log)
        db.commit()
        
        return {
            "message": "Master list uploaded successfully.",
            "added": added_count,
            "total_confirmed": total_count
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if os.path.exists(temp_path):
            os.remove(temp_path)

@router.post("/cross-reference")
def run_cross_reference(
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    # Fetch all confirmed numbers and normalize them just in case
    confirmed_records = db.query(ConfirmedPIO).all()
    confirmed_set = {record.phone_number.replace(" ", "").replace("-", "").lstrip('+') for record in confirmed_records}
    
    if not confirmed_set:
        return {"message": "No confirmed PIO numbers in the database.", "cases_flagged": 0}
        
    # Fetch all cases that have suspected PIO numbers
    cases = db.query(Case).filter(Case.suspected_pio_numbers.isnot(None), Case.suspected_pio_numbers != "").all()
    
    flagged_count = 0
    for case in cases:
        suspected_nums = [n.strip() for n in case.suspected_pio_numbers.split(",")]
        matches = []
        for n in suspected_nums:
            normalized_n = n.replace(" ", "").replace("-", "").lstrip('+')
            if normalized_n in confirmed_set:
                matches.append(n) # Keep original formatting for display
                
        if matches:
            case.has_confirmed_pio = True
            case.confirmed_pio_matches = ", ".join(matches)
            flagged_count += 1
        else:
            case.has_confirmed_pio = False
            case.confirmed_pio_matches = None
            
    db.commit()
    
    log = AuditLog(
        username=current_user.username,
        action="RUN_PIO_CROSS_REFERENCE",
        details=f"Ran cross-reference check. Flagged {flagged_count} cases."
    )
    db.add(log)
    db.commit()
    
    return {
        "message": "Cross-reference check completed.",
        "cases_flagged": flagged_count
    }
